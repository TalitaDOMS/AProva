from flask import Flask,render_template,request,redirect
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
ARQUIVO  = 'notas.xlsx'

# Criação do arquivo
if not os.path.exists(ARQUIVO):
    wb = Workbook()
    ws = wb.active
    ws.append(['Aluno','RA','Tarefas','Prova Mensal','Prova Bimestral','Média Final','Status'])
    ws['H1'] = 1
    wb.save(ARQUIVO)

# Paths do site
@app.route('/')
def index():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    dados = list(ws.iter_rows(min_row=2, max_col=7, values_only=True))
    return render_template('index.html', dados=dados)

@app.route('/cadastro')
def cadastro():
    return render_template('cadastro.html')

@app.route('/resumo')
def resumo():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    
    return render_template('resumo.html')

#Funções
@app.route('/enviar', methods=['POST'])
def enviar():
    nome = request.form['nome']
    
    tarefas = float(request.form['tarefas'])
    provaMensal = float(request.form['provaMensal'])
    provaBimestral = float(request.form['provaBimestral'])

    # Cálculo da média final
    mediaFinal = round(((tarefas + provaMensal + provaBimestral) / 3), 2)
    
    # Situação do aluno
    status = ''
    if mediaFinal >= 7:
        status = 'Aprovado 🟩'
    elif mediaFinal >= 5:
        status = 'Recuperação 🟨'
    else:
        status = 'Reprovado 🟥'

    # Salvar no arquivo Excel
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    ra = ws['H1'].value
    ws.append([nome, ra, tarefas, provaMensal, provaBimestral, mediaFinal, status])
    ra += 1
    ws['H1'] = ra
    wb.save(ARQUIVO)
    return render_template('resumo.html', nome=nome, ra=ra, tarefas=tarefas, provaMensal=provaMensal,
                           provaBimestral=provaBimestral, mediaFinal=mediaFinal, status=status)

# Iniciar
if __name__ == '__main__':
    app.run(debug=True)